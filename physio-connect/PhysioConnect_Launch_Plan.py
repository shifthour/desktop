#!/usr/bin/env python3
"""
PhysioConnect - Complete Launch Plan Generator
Launch Date: 30 March 2026 | Start: 2 March 2026 | 28 days
Teams:
  SRH — Operations (based in UK)
  RCB — Technical  (based in Bangalore)
Every task has exactly ONE owner: SRH or RCB
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

CAT_CLR = {
    "Legal & Company Formation":    PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "Regulatory & Compliance":      PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "Finance & Banking":            PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "💰 Payment: SRH → RCB":       PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    "Insurance":                    PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "Technical - Infrastructure":   PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
    "Technical - Backend":          PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),
    "Technical - Frontend":         PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid"),
    "Technical - Payments":         PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid"),
    "Technical - Testing & QA":     PatternFill(start_color="EFEBE9", end_color="EFEBE9", fill_type="solid"),
    "Design & Branding":            PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    "Content & Copywriting":        PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid"),
    "Physio Recruitment":           PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid"),
    "Marketing & Pre-Launch":       PatternFill(start_color="FBE9E7", end_color="FBE9E7", fill_type="solid"),
    "Operations & Support":         PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid"),
    "Launch Day":                   PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid"),
    "Post-Launch (Week 1)":         PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "Post-Launch (Week 2-4)":       PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid"),
    "Post-Launch (Month 2)":        PatternFill(start_color="81C784", end_color="81C784", fill_type="solid"),
    "Post-Launch (Month 3+)":       PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid"),
}

SRH_FONT = Font(name="Calibri", bold=True, color="1565C0", size=11)
RCB_FONT = Font(name="Calibri", bold=True, color="7B1FA2", size=11)

PRIORITY_FONTS = {
    "CRITICAL": Font(name="Calibri", bold=True, color="D32F2F", size=10),
    "HIGH":     Font(name="Calibri", bold=True, color="E65100", size=10),
    "MEDIUM":   Font(name="Calibri", color="1565C0", size=10),
    "LOW":      Font(name="Calibri", color="388E3C", size=10),
}

bdr = Border(
    left=Side(style='thin', color='BDBDBD'),
    right=Side(style='thin', color='BDBDBD'),
    top=Side(style='thin', color='BDBDBD'),
    bottom=Side(style='thin', color='BDBDBD')
)

# ============================================================
# SHEET 1 — MASTER LAUNCH PLAN
# ============================================================
ws = wb.active
ws.title = "Master Launch Plan"

ws.merge_cells('A1:J1')
ws['A1'] = "PHYSIOCONNECT — COMPLETE LAUNCH PLAN"
ws['A1'].font = Font(name="Calibri", bold=True, size=16, color="1B2A4A")
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:J2')
ws['A2'] = "🚀 Launch: 30 March 2026  |  UK Home Physiotherapy Marketplace  |  Today: 2 March 2026  |  28 Days to Launch"
ws['A2'].font = Font(name="Calibri", size=11, color="666666", italic=True)
ws['A2'].alignment = Alignment(horizontal='center')
ws.row_dimensions[2].height = 22

ws.merge_cells('A3:J3')
ws['A3'] = "🔵 SRH = Operations Team (UK) — Legal, Compliance, Recruitment, Marketing, Finance, Content, Ops    |    🟣 RCB = Technical Team (Bangalore) — Dev, Infra, Payments, QA, Design"
ws['A3'].font = Font(name="Calibri", size=10, color="333333", bold=True)
ws['A3'].alignment = Alignment(horizontal='center')
ws['A3'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
ws.row_dimensions[3].height = 25

headers = ["#", "Date", "Day", "Category", "Task", "Details / Notes", "Owner", "Priority", "Status", "Dependencies"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=col, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = bdr
ws.row_dimensions[5].height = 30

widths = [5, 14, 12, 28, 52, 58, 10, 12, 14, 25]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# TASK DATA — 28-day plan + post-launch through Month 3+
# (date, day_name, category, task, details, owner, priority, status, deps)
# owner is ONLY "SRH" or "RCB"
# ============================================================
T = [
    # ===================== DAY 1 — 2 Mar (Mon) =====================
    (date(2026,3,2), "Mon", "Legal & Company Formation",  "Register company with Companies House",          "Register Ltd company online (24hr turnaround). Cost: £12. Use web filing service",                                                          "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,2), "Mon", "Legal & Company Formation",  "Choose company name & check availability",       "Search Companies House register + domain availability. Finalise name",                                                                      "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,2), "Mon", "Legal & Company Formation",  "Register domain name (.co.uk + .com)",           "Namecheap / GoDaddy. Cost ~£10-20. Secure both extensions immediately",                                                                     "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,2), "Mon", "Legal & Company Formation",  "Appoint directors & shareholders",               "Form IN01. Decide share structure. Registered office address needed",                                                                       "SRH", "HIGH",     "Not Started", "Company name"),
    (date(2026,3,2), "Mon", "Finance & Banking",          "Apply for business bank account",                "Tide / Starling / Monzo Business (fastest approval). Need: Co. number, ID, proof of address",                                               "SRH", "CRITICAL", "Not Started", "Company reg"),
    (date(2026,3,2), "Mon", "Regulatory & Compliance",    "Research HCPC registration requirements",        "All physios must be HCPC registered. Understand verification process for platform",                                                         "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,2), "Mon", "Regulatory & Compliance",    "Research CSP membership verification",           "Chartered Society of Physiotherapy. Adds credibility. Plan how to verify",                                                                  "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,2), "Mon", "Technical - Infrastructure", "Set up GitHub repository & branching",           "Create repo, main/develop branches, .gitignore, README, CI basics",                                                                         "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,2), "Mon", "Technical - Infrastructure", "Set up development environment",                 "Next.js project (exists). Verify Node, install deps, lint config, env vars",                                                                "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,2), "Mon", "Design & Branding",          "Design logo and brand identity",                 "Colours, typography, brand guidelines. Fiverr/99designs for fast turnaround",                                                                "RCB", "HIGH",     "Not Started", ""),

    # ===================== DAY 2 — 3 Mar (Tue) =====================
    (date(2026,3,3), "Tue", "Legal & Company Formation",  "Draft Terms of Service",                         "Platform terms, booking terms, cancellation policy, liability. Template + solicitor review",                                                   "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,3), "Tue", "Legal & Company Formation",  "Draft Privacy Policy (UK GDPR)",                 "UK GDPR & Data Protection Act 2018. Data collected, processing, retention, rights",                                                         "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,3), "Tue", "Legal & Company Formation",  "Register with ICO",                              "Data protection registration. £35-40/yr. Legal requirement for personal data processing",                                                   "SRH", "CRITICAL", "Not Started", "Company reg"),
    (date(2026,3,3), "Tue", "Insurance",                  "Research & obtain Public Liability Insurance",   "Platform-level PLI. Hiscox / Simply Business / AXA. Require physios to have own PLI",                                                       "SRH", "CRITICAL", "Not Started", "Company reg"),
    (date(2026,3,3), "Tue", "Insurance",                  "Research Professional Indemnity Insurance",      "PI insurance for platform operations. Physios must have own PI via practice",                                                                "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,3), "Tue", "Technical - Infrastructure", "Set up cloud hosting (Vercel)",                  "Deploy Next.js to Vercel. Set up staging + production environments",                                                                        "RCB", "CRITICAL", "Not Started", "GitHub repo"),
    (date(2026,3,3), "Tue", "Technical - Infrastructure", "Set up database (Supabase/PostgreSQL)",          "Supabase (Postgres + Auth + Realtime). Tables, RLS policies. Free tier",                                                                    "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,3), "Tue", "Technical - Backend",        "Design complete database schema",                "Tables: users, physios, bookings, reviews, availability, payments, addresses, qualifications, services",                                    "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,3), "Tue", "Technical - Backend",        "Set up authentication system",                   "Supabase Auth or NextAuth.js. Email/password + Google OAuth. Separate patient & physio flows",                                              "RCB", "CRITICAL", "Not Started", "DB setup"),

    # ===================== DAY 3 — 4 Mar (Wed) — PAYMENT 1 =====================
    (date(2026,3,4), "Wed", "💰 Payment: SRH → RCB",     "SRH pays ₹50,000 (1st instalment) to RCB",      "1st of 3 instalments. Total ₹1,50,000. Bank transfer/UPI to RCB. Collect receipt immediately",                                              "SRH", "CRITICAL", "Not Started", "Bank account"),
    (date(2026,3,4), "Wed", "Legal & Company Formation",  "Register for HMRC / Corporation Tax",            "Register within 3 months. Get UTR number. Consider VAT threshold (£90k)",                                                                   "SRH", "HIGH",     "Not Started", "Company reg"),
    (date(2026,3,4), "Wed", "Legal & Company Formation",  "Draft Physiotherapist Agreement / Contract",     "Independent contractor agreement. Commission, standards, cancellation, insurance, HCPC",                                                     "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,4), "Wed", "Technical - Backend",        "Build physio registration & onboarding flow",    "Multi-step: personal details, qualifications, HCPC no., services, areas, availability",                                                     "RCB", "CRITICAL", "Not Started", "Auth system"),
    (date(2026,3,4), "Wed", "Technical - Backend",        "Build HCPC verification integration",            "HCPC online register - scraper/API to verify physio status. Critical for trust & safety",                                                   "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,4), "Wed", "Technical - Frontend",       "Build landing page / homepage",                  "Hero, how it works, benefits, trust signals, CTAs. Mobile-first design",                                                                    "RCB", "CRITICAL", "Not Started", "Design ready"),
    (date(2026,3,4), "Wed", "Content & Copywriting",      "Write website copy — Homepage",                  "Value proposition, how it works steps, patient benefits, trust messaging",                                                                   "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,4), "Wed", "Physio Recruitment",         "Create physio recruitment campaign plan",        "Channels: LinkedIn, CSP jobs board, FB groups, Indeed, local networks",                                                                     "SRH", "CRITICAL", "Not Started", ""),

    # ===================== DAY 4 — 5 Mar (Thu) — PAYMENT 2 =====================
    (date(2026,3,5), "Thu", "💰 Payment: SRH → RCB",     "SRH pays ₹50,000 (2nd instalment) to RCB",      "2nd instalment. Running total: ₹1,00,000 / ₹1,50,000. Bank transfer. Get receipt",                                                         "SRH", "CRITICAL", "Not Started", "1st paid"),
    (date(2026,3,5), "Thu", "Technical - Frontend",       "Build patient registration flow",                "Sign-up: name, email, phone, postcode, medical conditions (optional)",                                                                      "RCB", "HIGH",     "Not Started", "Auth system"),
    (date(2026,3,5), "Thu", "Technical - Backend",        "Build physio profile & dashboard",               "Profile: bio, qualifications, specialisms, reviews, availability calendar, earnings",                                                        "RCB", "CRITICAL", "Not Started", "Physio reg"),
    (date(2026,3,5), "Thu", "Technical - Backend",        "Build service listing & search engine",          "Services list (sports, rehab, elderly). Search by postcode + service type",                                                                  "RCB", "CRITICAL", "Not Started", "DB schema"),
    (date(2026,3,5), "Thu", "Technical - Backend",        "Build availability management system",           "Weekly availability, calendar view, block dates, recurring schedules",                                                                       "RCB", "HIGH",     "Not Started", "Physio dashboard"),
    (date(2026,3,5), "Thu", "Physio Recruitment",         "Post job listings on all channels",              "LinkedIn, Indeed, CSP job board, Gumtree, physio FB groups, Reed.co.uk",                                                                    "SRH", "CRITICAL", "Not Started", "Recruitment plan"),
    (date(2026,3,5), "Thu", "Physio Recruitment",         "Direct outreach to local physios",               "Find private physios in target area. Personal email/LinkedIn with value prop",                                                               "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,5), "Thu", "Content & Copywriting",      "Write copy — About page, How it Works",         "Detailed how-it-works for patients & physios. FAQ content drafts",                                                                          "SRH", "HIGH",     "Not Started", ""),

    # ===================== DAY 5 — 6 Mar (Fri) — PAYMENT 3 =====================
    (date(2026,3,6), "Fri", "💰 Payment: SRH → RCB",     "SRH pays ₹50,000 (3rd / FINAL instalment)",     "FINAL instalment. ₹1,50,000 FULLY PAID. Get final receipt. Confirm settlement in writing",                                                  "SRH", "CRITICAL", "Not Started", "2nd paid"),
    (date(2026,3,6), "Fri", "Technical - Frontend",       "Build physio search results page",               "Cards: photo, name, specialisms, rating, distance, price, next available. Filters & sort",                                                   "RCB", "CRITICAL", "Not Started", "Search backend"),
    (date(2026,3,6), "Fri", "Technical - Frontend",       "Build physio public profile page",               "Full profile: about, qualifications, reviews, services & prices, book-now CTA",                                                              "RCB", "HIGH",     "Not Started", "Profile backend"),
    (date(2026,3,6), "Fri", "Technical - Payments",       "Set up Stripe Connect account",                  "Apply for UK Stripe Connect. Marketplace model. Connected accounts for physios",                                                             "SRH", "CRITICAL", "Not Started", "Company + Bank"),
    (date(2026,3,6), "Fri", "Marketing & Pre-Launch",     "Set up all social media accounts",               "Instagram, Facebook Page, LinkedIn Company, Twitter/X. Consistent branding",                                                                 "SRH", "HIGH",     "Not Started", "Logo ready"),
    (date(2026,3,6), "Fri", "Design & Branding",          "Design social media templates",                  "Post templates, story templates, brand-consistent look. Canva / Figma",                                                                     "RCB", "MEDIUM",   "Not Started", "Branding"),

    # ===================== DAY 6 — 7 Mar (Sat) =====================
    (date(2026,3,7), "Sat", "Technical - Payments",       "Build payment flow (Stripe integration)",        "Patient pays → Stripe holds → appointment done → release to physio minus commission",                                                       "RCB", "CRITICAL", "Not Started", "Stripe account"),
    (date(2026,3,7), "Sat", "Technical - Backend",        "Build booking system",                           "Select service → date/time → address → confirm → pay → confirmation emails",                                                                "RCB", "CRITICAL", "Not Started", "Availability + Payments"),
    (date(2026,3,7), "Sat", "Technical - Backend",        "Build email notification system",                "Booking confirm, 24hr reminders, cancellation notices, welcome. SendGrid / Resend",                                                         "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,7), "Sat", "Marketing & Pre-Launch",     "Create Google My Business listing",              "GMB for local SEO. Business details, photos, services",                                                                                     "SRH", "MEDIUM",   "Not Started", "Company reg"),

    # ===================== DAY 7 — 8 Mar (Sun) =====================
    (date(2026,3,8), "Sun", "Technical - Backend",        "Build booking management — physio side",         "Upcoming bookings, accept/decline, patient details, navigate to address, mark complete",                                                     "RCB", "CRITICAL", "Not Started", "Booking system"),
    (date(2026,3,8), "Sun", "Technical - Backend",        "Build booking management — patient side",        "View bookings, upcoming & past, cancel/reschedule, leave review",                                                                           "RCB", "HIGH",     "Not Started", "Booking system"),
    (date(2026,3,8), "Sun", "Technical - Frontend",       "Build booking flow UI (wizard)",                 "Step-by-step booking wizard. Mobile-responsive. Clean, reassuring design",                                                                  "RCB", "CRITICAL", "Not Started", "Booking backend"),
    (date(2026,3,8), "Sun", "Content & Copywriting",      "Write physio onboarding guide",                  "Step-by-step: set up profile, manage bookings, get paid, support contacts",                                                                  "SRH", "HIGH",     "Not Started", ""),

    # ===================== DAY 8 — 9 Mar (Mon) =====================
    (date(2026,3,9), "Mon", "Technical - Backend",        "Build review & rating system",                   "Patient rates physio (1-5 stars) + written review. Physio can respond",                                                                      "RCB", "HIGH",     "Not Started", "Booking complete"),
    (date(2026,3,9), "Mon", "Technical - Backend",        "Build admin dashboard",                          "All bookings, users, physios, revenue, disputes, pending verifications. KPIs",                                                               "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,9), "Mon", "Technical - Frontend",       "Build patient dashboard",                        "My bookings, upcoming, past treatments, saved physios, profile settings",                                                                    "RCB", "HIGH",     "Not Started", "Booking mgmt"),
    (date(2026,3,9), "Mon", "Physio Recruitment",         "Follow up with interested physios",              "Respond to applications, onboarding calls, answer commission / platform questions",                                                          "SRH", "CRITICAL", "Not Started", "Listings posted"),
    (date(2026,3,9), "Mon", "Legal & Company Formation",  "Finalise platform commission structure",         "Decide commission rate (15-25%). Clear pricing model for physio agreement",                                                                  "SRH", "CRITICAL", "Not Started", ""),

    # ===================== DAY 9 — 10 Mar (Tue) =====================
    (date(2026,3,10), "Tue", "Technical - Backend",       "Build cancellation & refund system",             "24hr free cancel. Late cancel fee 50%. No-show handling. Auto Stripe refunds",                                                               "RCB", "CRITICAL", "Not Started", "Payment system"),
    (date(2026,3,10), "Tue", "Technical - Backend",       "Build postcode / location matching",             "postcodes.io (free UK API). Distance calc. Match physios within X miles",                                                                    "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,10), "Tue", "Technical - Backend",       "Build SMS notifications (Twilio)",               "Booking confirms & reminders via SMS. UK mobiles. Optional at launch",                                                                       "RCB", "MEDIUM",   "Not Started", ""),
    (date(2026,3,10), "Tue", "Physio Recruitment",        "Onboard first physios onto platform",            "Help create profiles, upload docs, set availability, verify HCPC",                                                                           "SRH", "CRITICAL", "Not Started", "Physio reg flow"),
    (date(2026,3,10), "Tue", "Operations & Support",      "Set up customer support email",                  "support@domain. Shared inbox: Zendesk free / Freshdesk / Gmail",                                                                             "SRH", "HIGH",     "Not Started", "Domain"),

    # ===================== DAY 10 — 11 Mar (Wed) =====================
    (date(2026,3,11), "Wed", "Technical - Frontend",      "Mobile responsiveness pass — all pages",         "Test & fix every page on mobile. iPhone & Android. Most bookings from mobile",                                                                "RCB", "CRITICAL", "Not Started", "All frontend"),
    (date(2026,3,11), "Wed", "Technical - Backend",       "Build in-app messaging system",                  "Secure messaging patient ↔ physio. Pre-appointment Qs. No phone sharing",                                                                    "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,11), "Wed", "Technical - Backend",       "Implement DBS check verification workflow",      "Enhanced DBS for home visits. Upload & admin verification flow",                                                                             "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,11), "Wed", "Regulatory & Compliance",   "Create safeguarding policy",                     "Safeguarding vulnerable adults. Required for home healthcare. DBS reqs",                                                                     "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,11), "Wed", "Regulatory & Compliance",   "Create complaints procedure",                    "Formal complaints handling. Align with HCPC standards. Escalation path",                                                                     "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,11), "Wed", "Content & Copywriting",     "Write FAQ page content",                         "How it works, pricing, insurance, qualifications, cancellation, safety",                                                                     "SRH", "HIGH",     "Not Started", ""),

    # ===================== DAY 11 — 12 Mar (Thu) =====================
    (date(2026,3,12), "Thu", "Technical - Backend",       "Build automated email sequences",                "Welcome series, post-booking follow-up, review request (24hr), re-engagement",                                                               "RCB", "HIGH",     "Not Started", "Email system"),
    (date(2026,3,12), "Thu", "Marketing & Pre-Launch",    "Set up Google Analytics & tracking",             "GA4, Meta Pixel, conversion tracking. Track: sign-ups, bookings, revenue",                                                                   "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,12), "Thu", "Marketing & Pre-Launch",    "Create pre-launch landing page with waitlist",   "Coming-soon page + email capture. Value prop, launch date countdown",                                                                        "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,12), "Thu", "Physio Recruitment",        "Continue onboarding — target 10-15 physios",     "Minimum viable supply. Focus 2-3 London boroughs or target city",                                                                           "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,12), "Thu", "Insurance",                 "Finalise & purchase all insurance policies",     "PLI + PI purchased. Certificates saved. Renewal dates diarised",                                                                             "SRH", "HIGH",     "Not Started", "Insurance research"),

    # ===================== DAY 12 — 13 Mar (Fri) =====================
    (date(2026,3,13), "Fri", "Technical - Testing & QA",  "Full E2E testing — Patient journey",             "Sign up → search → profile → book → pay → confirm → cancel → refund",                                                                      "RCB", "CRITICAL", "Not Started", "All features"),
    (date(2026,3,13), "Fri", "Technical - Testing & QA",  "Full E2E testing — Physio journey",              "Register → verify → availability → booking → navigate → complete → paid",                                                                   "RCB", "CRITICAL", "Not Started", "All features"),
    (date(2026,3,13), "Fri", "Technical - Testing & QA",  "Payment testing (Stripe test mode)",             "All scenarios: success, failed card, refund, physio payout, partial",                                                                        "RCB", "CRITICAL", "Not Started", "Payments"),
    (date(2026,3,13), "Fri", "Technical - Testing & QA",  "Basic security audit",                           "Auth flows, SQL injection, XSS, CSRF, rate limiting. OWASP top 10",                                                                         "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,13), "Fri", "Marketing & Pre-Launch",    "Create launch marketing materials",              "Social posts (2 weeks), email announcement, press release draft",                                                                            "SRH", "HIGH",     "Not Started", "Branding"),
    (date(2026,3,13), "Fri", "Marketing & Pre-Launch",    "Set up email marketing (Mailchimp/Brevo)",       "Account, import waitlist, welcome email series, booking reminders",                                                                          "SRH", "HIGH",     "Not Started", ""),

    # ===================== DAY 13 — 14 Mar (Sat) =====================
    (date(2026,3,14), "Sat", "Technical - Testing & QA",  "Fix all bugs from E2E testing",                  "Address every critical & high bug found. Regression test after fixes",                                                                       "RCB", "CRITICAL", "Not Started", "E2E testing"),
    (date(2026,3,14), "Sat", "Technical - Infrastructure","Set up error monitoring (Sentry)",               "Frontend & backend error tracking. Slack/email alerts for critical failures",                                                                "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,14), "Sat", "Content & Copywriting",     "Write blog posts for launch",                    "2-3 SEO articles: 'Benefits of home physio', 'How to choose a physio'",                                                                     "SRH", "MEDIUM",   "Not Started", ""),
    (date(2026,3,14), "Sat", "Operations & Support",      "Create support documentation / help centre",     "How to book, payment Qs, cancellation, physio FAQ, patient FAQ",                                                                             "SRH", "HIGH",     "Not Started", ""),

    # ===================== DAY 14 — 15 Mar (Sun) =====================
    (date(2026,3,15), "Sun", "Technical - Frontend",      "Performance optimisation",                       "Lighthouse audit. Image opt, lazy loading, code splitting. Target 90+",                                                                      "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,15), "Sun", "Technical - Frontend",      "SEO setup",                                      "Meta tags, OG, Schema.org LocalBusiness, sitemap.xml, robots.txt",                                                                          "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,15), "Sun", "Operations & Support",      "Create internal operations playbook",            "Disputes, refunds, complaints, emergencies, escalation procedures",                                                                          "SRH", "HIGH",     "Not Started", ""),

    # ===================== DAY 15 — 16 Mar (Mon) =====================
    (date(2026,3,16), "Mon", "Technical - Infrastructure", "Set up custom domain & SSL",                    "Point domain → Vercel. SSL auto. www redirect. Email forwarding",                                                                            "RCB", "CRITICAL", "Not Started", "Domain + Hosting"),
    (date(2026,3,16), "Mon", "Technical - Infrastructure", "Set up production database & backups",          "Production Supabase. Automated daily backups. DR plan documented",                                                                           "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,16), "Mon", "Technical - Infrastructure", "Set up CDN & image hosting",                    "Cloudflare / Vercel Edge. Cloudinary or S3 for photos & docs",                                                                               "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,16), "Mon", "Technical - Payments",       "Switch Stripe to live mode",                    "Activate live. Verify bank connected. Real £1 test. Webhook endpoints",                                                                      "RCB", "CRITICAL", "Not Started", "Stripe test pass"),
    (date(2026,3,16), "Mon", "Physio Recruitment",         "Verify all onboarded physios",                  "HCPC ✓, insurance docs ✓, DBS ✓, profile complete ✓, bank for payout ✓",                                                                    "SRH", "CRITICAL", "Not Started", "Onboarding"),
    (date(2026,3,16), "Mon", "Marketing & Pre-Launch",     "Set up Google Ads account (prep)",              "Account, conversion tracking, draft campaign. Postcode + physio keywords",                                                                   "SRH", "MEDIUM",   "Not Started", ""),

    # ===================== DAY 16 — 17 Mar (Tue) =====================
    (date(2026,3,17), "Tue", "Technical - Testing & QA",  "Staging environment full test",                  "Deploy to staging. Complete walkthrough all flows. Test with real physio accounts",                                                           "RCB", "CRITICAL", "Not Started", "All dev done"),
    (date(2026,3,17), "Tue", "Technical - Testing & QA",  "Cross-browser testing",                          "Chrome, Safari, Firefox, Edge. Mobile Safari (iPhone), Chrome (Android)",                                                                    "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,17), "Tue", "Technical - Testing & QA",  "Basic load testing",                             "Simulate 50-100 concurrent users. Identify bottlenecks. Booking under load",                                                                 "RCB", "MEDIUM",   "Not Started", ""),
    (date(2026,3,17), "Tue", "Physio Recruitment",        "Physio platform walkthrough session",            "Video call / webinar with all physios. Demo platform. Q&A. Gather feedback",                                                                 "SRH", "HIGH",     "Not Started", "Physios onboarded"),

    # ===================== DAY 17 — 18 Mar (Wed) =====================
    (date(2026,3,18), "Wed", "Technical - Testing & QA",  "Fix all bugs from staging / cross-browser",     "Address every remaining issue. Regression test",                                                                                             "RCB", "CRITICAL", "Not Started", "Staging test"),
    (date(2026,3,18), "Wed", "Technical - Infrastructure", "Production deployment (soft launch)",           "Deploy to production. Verify all services. Monitor errors. Invite-only mode",                                                                "RCB", "CRITICAL", "Not Started", "All tests pass"),
    (date(2026,3,18), "Wed", "Technical - Infrastructure", "Set up uptime monitoring",                      "UptimeRobot / Better Uptime. Monitor: site, API, DB, payments. Alerts",                                                                      "RCB", "HIGH",     "Not Started", "Prod deploy"),
    (date(2026,3,18), "Wed", "Marketing & Pre-Launch",     "Schedule pre-launch social media (10 days)",    "Daily teaser content. Behind the scenes, physio spotlights, countdown",                                                                      "SRH", "HIGH",     "Not Started", "Social accounts"),

    # ===================== DAY 18 — 19 Mar (Thu) =====================
    (date(2026,3,19), "Thu", "Technical - Payments",       "Verify live payment flow end-to-end",           "Real £1 test booking full flow. Verify physio payout. All emails sent correctly",                                                             "RCB", "CRITICAL", "Not Started", "Live Stripe"),
    (date(2026,3,19), "Thu", "Marketing & Pre-Launch",     "Send email to waitlist — launch coming",        "Email all pre-registered users. Early-bird offer. Launch date reveal",                                                                        "SRH", "HIGH",     "Not Started", "Email mktg"),
    (date(2026,3,19), "Thu", "Marketing & Pre-Launch",     "Prepare press release",                         "Local press, health publications, physio trade press. Draft newsworthy angle",                                                                "SRH", "MEDIUM",   "Not Started", ""),
    (date(2026,3,19), "Thu", "Operations & Support",       "Set up live chat widget",                       "Tawk.to (free) or Crisp. Key pages: booking, contact, physio sign-up",                                                                       "RCB", "MEDIUM",   "Not Started", ""),

    # ===================== DAY 19-20 — 20-21 Mar (Fri-Sat) =====================
    (date(2026,3,20), "Fri", "Technical - Testing & QA",  "Beta test with 5-10 friendly users",            "Real users make real test bookings. Feedback on UX, flow, clarity, trust",                                                                    "RCB", "CRITICAL", "Not Started", "Soft launch"),
    (date(2026,3,20), "Fri", "Physio Recruitment",        "Confirm all physios ready for launch",           "Final check: profiles ✓, availability ✓, insurance ✓, HCPC ✓, payout ✓",                                                                   "SRH", "CRITICAL", "Not Started", ""),
    (date(2026,3,21), "Sat", "Technical - Testing & QA",  "Fix critical issues from beta feedback",         "Showstopper bugs or UX issues. Fast turnaround fixes",                                                                                       "RCB", "CRITICAL", "Not Started", "Beta test"),
    (date(2026,3,21), "Sat", "Marketing & Pre-Launch",    "Pre-launch social media campaign kicks off",     "Daily countdown posts. Physio spotlights. Beta testimonials",                                                                                "SRH", "HIGH",     "Not Started", "Content ready"),

    # ===================== DAY 21 — 22 Mar (Sun) =====================
    (date(2026,3,22), "Sun", "Technical - Testing & QA",  "Round 2 — re-test all fixed issues",            "Verify every beta fix. Full regression on critical paths",                                                                                   "RCB", "HIGH",     "Not Started", "Beta fixes"),
    (date(2026,3,22), "Sun", "Content & Copywriting",     "Write launch-week email sequence",              "Day-of launch, day+1 follow-up, day+3 reminder, day+7 re-engage",                                                                           "SRH", "HIGH",     "Not Started", ""),

    # ===================== DAY 22 — 23 Mar (Mon) =====================
    (date(2026,3,23), "Mon", "Legal & Company Formation",  "Final legal review — all policies & pages",    "Solicitor reviews T&Cs, Privacy, Physio Agreement. GDPR sign-off",                                                                           "SRH", "CRITICAL", "Not Started", "All policies"),
    (date(2026,3,23), "Mon", "Technical - Infrastructure", "Final production deployment",                  "Deploy final version. DB migrations. Clear caches. Verify integrations",                                                                     "RCB", "CRITICAL", "Not Started", "All fixes"),
    (date(2026,3,23), "Mon", "Technical - Testing & QA",   "Smoke test production",                        "Quick test critical paths on live: sign up, search, book, pay, confirm",                                                                     "RCB", "CRITICAL", "Not Started", "Final deploy"),

    # ===================== DAY 23 — 24 Mar (Tue) =====================
    (date(2026,3,24), "Tue", "Operations & Support",      "Brief support process / prepare for launch",    "FAQ responses, common issue resolutions, escalation contacts, shift rota",                                                                   "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,24), "Tue", "Marketing & Pre-Launch",    "Send press release to media contacts",          "Local press, health publications, physio outlets. Follow up with key journos",                                                                "SRH", "MEDIUM",   "Not Started", "Press release"),
    (date(2026,3,24), "Tue", "Marketing & Pre-Launch",    "Prepare launch-day social media content",       "Announcement posts all platforms. Physio spotlight threads. Special offers",                                                                  "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,24), "Tue", "Technical - Infrastructure", "Performance & stress test production",         "Final load test on prod. Verify auto-scaling. Check response times under load",                                                               "RCB", "HIGH",     "Not Started", ""),

    # ===================== DAY 24-25 — 25-26 Mar (Wed-Thu) =====================
    (date(2026,3,25), "Wed", "Technical - Testing & QA",  "Final bug fixes & polish",                      "Last-minute fixes. UI polish. Copy corrections. Final QA pass",                                                                              "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,25), "Wed", "Legal & Company Formation", "Ensure all legal pages live on website",        "T&Cs, Privacy, Cookie Policy, Complaints — all linked in footer and working",                                                                "SRH", "CRITICAL", "Not Started", "Legal review"),
    (date(2026,3,26), "Thu", "Technical - Infrastructure","Pre-launch checklist completion",               "SSL ✓, backups ✓, monitoring ✓, Sentry ✓, Stripe live ✓, emails ✓, all pages ✓",                                                             "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,26), "Thu", "Technical - Infrastructure","Scale infrastructure if needed",                "Check Vercel plan, Supabase limits, Stripe limits, email limits. Upgrade if needed",                                                         "RCB", "HIGH",     "Not Started", ""),
    (date(2026,3,26), "Thu", "Operations & Support",      "Notify all physios — launch in 4 days",        "Personal message to each. Confirm launch-week availability. Build excitement",                                                                "SRH", "HIGH",     "Not Started", ""),

    # ===================== DAY 26-27 — 27-28 Mar (Fri-Sat) =====================
    (date(2026,3,27), "Fri", "Marketing & Pre-Launch",    "Schedule all launch-day posts & emails",        "Pre-schedule social posts for 30 March. Email blast ready. Ads ready to activate",                                                            "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,27), "Fri", "Technical - Testing & QA",  "Final smoke test — everything green",           "Last check: all critical paths, payments, emails, monitoring. Sign-off from RCB",                                                             "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,3,28), "Sat", "Operations & Support",      "Dry run — simulate launch day operations",     "Run through launch-day checklist. Test support flow. Test monitoring alerts",                                                                  "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,28), "Sat", "Marketing & Pre-Launch",    "Final countdown posts — 2 days to go",         "Excitement building. Behind-the-scenes. Physio quotes. Launch teaser",                                                                       "SRH", "HIGH",     "Not Started", ""),

    # ===================== DAY 28 — 29 Mar (Sun) — PRE-LAUNCH =====================
    (date(2026,3,29), "Sun", "Operations & Support",      "Final call with all physios — launch tomorrow", "Last sync. Confirm everyone is ready. Share launch-day timeline",                                                                            "SRH", "HIGH",     "Not Started", ""),
    (date(2026,3,29), "Sun", "Technical - Infrastructure","Final system health check",                     "All services green. Backups ran. Monitoring active. On-call schedule confirmed",                                                               "RCB", "CRITICAL", "Not Started", ""),

    # ===================== 🚀 DAY 29 — 30 Mar (Mon) — LAUNCH DAY =====================
    (date(2026,3,30), "Mon", "Launch Day", "🚀 LAUNCH — Remove invite-only mode, GO LIVE",              "Open platform to public. Flip the switch. Remove beta flags. THIS IS IT!",                                                                    "RCB", "CRITICAL", "Not Started", "All pre-launch"),
    (date(2026,3,30), "Mon", "Launch Day", "Monitor systems closely (ALL DAY)",                          "Error rates, server load, payments, email delivery. Hotfix ready. RCB on-call",                                                               "RCB", "CRITICAL", "Not Started", "Launch"),
    (date(2026,3,30), "Mon", "Launch Day", "Publish launch social media posts — all platforms",          "LinkedIn, Instagram, Facebook, Twitter/X. Engage with comments all day",                                                                     "SRH", "CRITICAL", "Not Started", "Launch"),
    (date(2026,3,30), "Mon", "Launch Day", "Send launch email to full mailing list",                     "Launch announcement + offer/promotion. Clear CTA: book your first appointment",                                                               "SRH", "CRITICAL", "Not Started", "Launch"),
    (date(2026,3,30), "Mon", "Launch Day", "Activate Google Ads campaign",                               "Turn on PPC: '[city] home physio', 'physio home visit near me'",                                                                             "SRH", "HIGH",     "Not Started", "Ads setup"),
    (date(2026,3,30), "Mon", "Launch Day", "Monitor first real bookings personally",                     "Ensure first bookings go smoothly. Call physio & patient. Be on standby",                                                                     "SRH", "CRITICAL", "Not Started", "Launch"),
    (date(2026,3,30), "Mon", "Launch Day", "Respond to ALL enquiries same day",                          "Fast response critical. Support email, live chat, social DMs. SRH handles all",                                                               "SRH", "CRITICAL", "Not Started", "Launch"),

    # ================================================================
    # POST-LAUNCH — WEEK 1 (31 Mar – 5 Apr)
    # ================================================================
    (date(2026,3,31), "Tue", "Post-Launch (Week 1)", "Review Day 1 metrics & fix issues",               "Sign-ups, bookings, revenue, errors, feedback. Hotfix any critical bugs",                                                                      "RCB", "CRITICAL", "Not Started", "Launch done"),
    (date(2026,3,31), "Tue", "Post-Launch (Week 1)", "Day 1 ops review — what worked, what broke",      "SRH reviews: support volume, physio feedback, patient complaints, booking flow",                                                               "SRH", "CRITICAL", "Not Started", "Launch done"),
    (date(2026,4,1),  "Wed", "Post-Launch (Week 1)", "Gather feedback from first patients & physios",   "Call/message first users. NPS score. What worked? Pain points? Quick wins",                                                                    "SRH", "HIGH",     "Not Started", ""),
    (date(2026,4,1),  "Wed", "Post-Launch (Week 1)", "Iterate — quick-win UX fixes",                    "Copy changes, flow improvements, UI tweaks based on real user behaviour",                                                                     "RCB", "HIGH",     "Not Started", "Feedback"),
    (date(2026,4,2),  "Thu", "Post-Launch (Week 1)", "Continue physio recruitment — expand areas",       "Target: 25-30 physios month 1. Start recruiting in adjacent postcodes",                                                                       "SRH", "HIGH",     "Not Started", ""),
    (date(2026,4,2),  "Thu", "Post-Launch (Week 1)", "Run first Google Ads optimisation",               "Review ad performance, pause low performers, optimise bids, negative keywords",                                                                "SRH", "HIGH",     "Not Started", "Ads running"),
    (date(2026,4,3),  "Fri", "Post-Launch (Week 1)", "Verify first physio payout cycle",                "Confirm Stripe payouts reaching physios. Resolve any payment issues ASAP",                                                                     "RCB", "CRITICAL", "Not Started", ""),
    (date(2026,4,3),  "Fri", "Post-Launch (Week 1)", "Set up weekly metrics dashboard",                 "GMV, bookings/week, new users, new physios, retention, NPS, commission revenue",                                                               "RCB", "HIGH",     "Not Started", ""),
    (date(2026,4,5),  "Sun", "Post-Launch (Week 1)", "Week 1 review & retrospective",                   "Full review with SRH + RCB. What went well. What to improve. Month 1 goals",                                                                  "SRH", "HIGH",     "Not Started", ""),

    # ================================================================
    # POST-LAUNCH — WEEKS 2-4 (6 Apr – 26 Apr)
    # ================================================================
    (date(2026,4,6),  "Mon", "Post-Launch (Week 2-4)", "Build recurring / follow-up booking feature",    "Patients book weekly sessions. Package discounts. Auto-rebook",                                                                                "RCB", "HIGH",     "Not Started", ""),
    (date(2026,4,6),  "Mon", "Post-Launch (Week 2-4)", "Launch referral programme",                      "Patient refers friend → both get £10 off. Physio referral bonus. Track referrals",                                                             "SRH", "HIGH",     "Not Started", ""),
    (date(2026,4,8),  "Wed", "Post-Launch (Week 2-4)", "Expand to 2nd geographic area",                  "Based on demand data. Recruit physios in new area. Local marketing push",                                                                     "SRH", "HIGH",     "Not Started", ""),
    (date(2026,4,8),  "Wed", "Post-Launch (Week 2-4)", "Build automated review-request emails",          "24hr post-appointment → request review. Follow up day 3 if no review",                                                                        "RCB", "MEDIUM",   "Not Started", ""),
    (date(2026,4,10), "Fri", "Post-Launch (Week 2-4)", "Establish GP surgery partnerships",              "Leaflets in waiting rooms. Referral pathway from GPs. Outreach to local practices",                                                            "SRH", "HIGH",     "Not Started", ""),
    (date(2026,4,10), "Fri", "Post-Launch (Week 2-4)", "Build physio earnings & tax report",             "Monthly/annual earnings summary. CSV download. Help with self-assessment",                                                                     "RCB", "HIGH",     "Not Started", ""),
    (date(2026,4,13), "Mon", "Post-Launch (Week 2-4)", "Content marketing — launch weekly blog",         "SEO: 'exercises for back pain', 'when to see a physio'. Weekly cadence",                                                                       "SRH", "MEDIUM",   "Not Started", ""),
    (date(2026,4,13), "Mon", "Post-Launch (Week 2-4)", "Build advanced search filters",                  "Specialisation, gender, language, price range, availability, rating filters",                                                                  "RCB", "HIGH",     "Not Started", ""),
    (date(2026,4,15), "Wed", "Post-Launch (Week 2-4)", "Run first Facebook / Instagram ad campaign",     "Target: local area, 25-65, health-interested. A/B test creatives",                                                                             "SRH", "HIGH",     "Not Started", ""),
    (date(2026,4,15), "Wed", "Post-Launch (Week 2-4)", "Build patient medical notes (physio-side)",      "Secure clinical notes per patient. Treatment plans. SOAP notes. GDPR storage",                                                                 "RCB", "HIGH",     "Not Started", ""),
    (date(2026,4,17), "Fri", "Post-Launch (Week 2-4)", "Research NHS referral pathway integration",      "Can GPs refer via platform? NHS Choose & Book. Private + NHS hybrid?",                                                                         "SRH", "MEDIUM",   "Not Started", ""),
    (date(2026,4,17), "Fri", "Post-Launch (Week 2-4)", "Implement push notifications (web)",             "Browser push: new bookings (physio), confirms (patient), reminders",                                                                           "RCB", "MEDIUM",   "Not Started", ""),
    (date(2026,4,20), "Mon", "Post-Launch (Week 2-4)", "Month 1 board review / investor update",         "Metrics: users, bookings, revenue, growth, unit economics. Plan Month 2",                                                                      "SRH", "HIGH",     "Not Started", ""),
    (date(2026,4,20), "Mon", "Post-Launch (Week 2-4)", "Apply for startup grants / funding (if needed)", "Innovate UK, SEIS/EIS angels, Seedrs/Crowdcube. Start Fund, Tech Nation",                                                                      "SRH", "MEDIUM",   "Not Started", ""),
    (date(2026,4,22), "Wed", "Post-Launch (Week 2-4)", "File first VAT return (if applicable)",          "If approaching £90k threshold. Register. MTD-compliant software",                                                                              "SRH", "MEDIUM",   "Not Started", ""),
    (date(2026,4,22), "Wed", "Post-Launch (Week 2-4)", "Hire first customer support person (part-time)", "Handle growing enquiries. Train on playbook. Cover evenings/weekends",                                                                         "SRH", "MEDIUM",   "Not Started", ""),

    # ================================================================
    # POST-LAUNCH — MONTH 2 (Late Apr – Late May)
    # ================================================================
    (date(2026,4,27), "Mon", "Post-Launch (Month 2)", "Begin mobile app development (React Native)",     "Native patient app. Push notifs, location, quick rebook. MVP scope",                                                                           "RCB", "HIGH",     "Not Started", ""),
    (date(2026,4,27), "Mon", "Post-Launch (Month 2)", "Automate DBS check verification",                "Provider integration (uCheck / DBS Online). Reduce manual verification",                                                                      "RCB", "HIGH",     "Not Started", ""),
    (date(2026,5,1),  "Fri", "Post-Launch (Month 2)", "Build physio mobile app / PWA",                  "Physios need mobile: bookings, directions, availability on-the-go",                                                                            "RCB", "HIGH",     "Not Started", ""),
    (date(2026,5,1),  "Fri", "Post-Launch (Month 2)", "Launch corporate wellness partnerships",          "Approach companies for employee physio benefits. B2B pricing. Corp accounts",                                                                  "SRH", "HIGH",     "Not Started", ""),
    (date(2026,5,4),  "Mon", "Post-Launch (Month 2)", "Implement insurance claim assistance",            "Help patients claim on private insurance (Bupa, AXA, Vitality). Receipt gen",                                                                  "RCB", "HIGH",     "Not Started", ""),
    (date(2026,5,4),  "Mon", "Post-Launch (Month 2)", "Expand to 3rd & 4th geographic areas",            "Data-driven expansion. Recruit physios. Localised marketing per area",                                                                         "SRH", "HIGH",     "Not Started", ""),
    (date(2026,5,8),  "Fri", "Post-Launch (Month 2)", "Build appointment reminder automation",           "SMS + email: 24hr + 2hr before. Reduce no-shows by 40%+",                                                                                     "RCB", "HIGH",     "Not Started", ""),
    (date(2026,5,8),  "Fri", "Post-Launch (Month 2)", "Set up physio ambassador programme",              "Top physios as ambassadors. Recruit other physios for bonus. Brand advocates",                                                                  "SRH", "MEDIUM",   "Not Started", ""),
    (date(2026,5,11), "Mon", "Post-Launch (Month 2)", "Implement physio performance scoring",            "Internal score: reviews, on-time rate, completion rate, retention",                                                                             "RCB", "MEDIUM",   "Not Started", ""),
    (date(2026,5,11), "Mon", "Post-Launch (Month 2)", "Hire operations assistant (part-time)",            "Physio vetting, customer Qs, day-to-day ops. Free up founder time",                                                                            "SRH", "MEDIUM",   "Not Started", ""),
    (date(2026,5,15), "Fri", "Post-Launch (Month 2)", "First quarterly compliance audit",                "GDPR data audit, HCPC compliance, insurance renewals, policy review",                                                                          "SRH", "HIGH",     "Not Started", ""),
    (date(2026,5,18), "Mon", "Post-Launch (Month 2)", "Month 2 review & growth planning",               "Full metrics. Growth rate, retention, LTV, CAC. Set Month 3 targets",                                                                          "SRH", "HIGH",     "Not Started", ""),

    # ================================================================
    # POST-LAUNCH — MONTH 3+ (Late May onwards)
    # ================================================================
    (date(2026,5,25), "Mon", "Post-Launch (Month 3+)", "Launch patient mobile app (App Store + Play Store)", "Submit Apple + Google. ASO. Launch marketing campaign for app",                                                                             "RCB", "HIGH",     "Not Started", "App dev done"),
    (date(2026,5,25), "Mon", "Post-Launch (Month 3+)", "Explore seed / Series fundraising",                 "Pitch deck, financial model, investor outreach. Target £250k-500k seed",                                                                    "SRH", "HIGH",     "Not Started", "Month 2 metrics"),
    (date(2026,6,1),  "Mon", "Post-Launch (Month 3+)", "Implement video consultation feature",              "Pre-assessment video calls. Follow-up consults. Expand beyond home visits",                                                                 "RCB", "MEDIUM",   "Not Started", ""),
    (date(2026,6,1),  "Mon", "Post-Launch (Month 3+)", "National expansion planning",                       "Roadmap: top 10 UK cities. Recruitment pipeline. Marketing per city",                                                                       "SRH", "HIGH",     "Not Started", ""),
    (date(2026,6,8),  "Mon", "Post-Launch (Month 3+)", "Build multi-language support",                      "English, Polish, Hindi, Bengali, Urdu (UK demographics). i18n framework",                                                                   "RCB", "MEDIUM",   "Not Started", ""),
    (date(2026,6,8),  "Mon", "Post-Launch (Month 3+)", "Establish Advisory Board (clinical governance)",     "Senior physio, healthcare lawyer, NHS advisor, tech advisor. Quarterly meets",                                                               "SRH", "HIGH",     "Not Started", ""),
    (date(2026,6,15), "Mon", "Post-Launch (Month 3+)", "Implement AI-powered physio matching",               "Match by condition, location, specialisation, reviews, price, availability",                                                                "RCB", "MEDIUM",   "Not Started", ""),
    (date(2026,6,15), "Mon", "Post-Launch (Month 3+)", "Launch loyalty / subscription programme",            "Monthly sub: 2 sessions at discount. Annual & family plans",                                                                               "SRH", "HIGH",     "Not Started", ""),
    (date(2026,6,22), "Mon", "Post-Launch (Month 3+)", "Build API for clinic / hospital integrations",       "Clinics refer patients, check availability, track outcomes. B2B growth",                                                                    "RCB", "MEDIUM",   "Not Started", ""),
    (date(2026,6,22), "Mon", "Post-Launch (Month 3+)", "CQC registration assessment (if applicable)",        "Care Quality Commission. Assess requirements. Begin registration if needed",                                                                "SRH", "HIGH",     "Not Started", ""),
    (date(2026,6,30), "Tue", "Post-Launch (Month 3+)", "Quarter 1 full business review",                     "Revenue, growth, burn, metrics, NPS, competition. Set Q2 OKRs",                                                                            "SRH", "HIGH",     "Not Started", ""),
]

# ── Write tasks ──
for idx, t in enumerate(T, 1):
    r = idx + 5
    dt, day, cat, task, det, own, pri, st, dep = t
    ws.cell(r,1,idx).alignment = Alignment(horizontal='center')
    ws.cell(r,2,dt.strftime("%d %b %Y")).alignment = Alignment(horizontal='center')
    ws.cell(r,3,day).alignment = Alignment(horizontal='center')
    ws.cell(r,4,cat)
    ws.cell(r,5,task)
    ws.cell(r,6,det).alignment = Alignment(wrap_text=True)
    oc = ws.cell(r,7,own); oc.alignment = Alignment(horizontal='center')
    oc.font = SRH_FONT if own=="SRH" else RCB_FONT
    ws.cell(r,8,pri)
    ws.cell(r,9,st).alignment = Alignment(horizontal='center')
    ws.cell(r,10,dep).alignment = Alignment(wrap_text=True)
    if cat in CAT_CLR:
        for c in range(1,11): ws.cell(r,c).fill = CAT_CLR[cat]
    if pri in PRIORITY_FONTS: ws.cell(r,8).font = PRIORITY_FONTS[pri]
    oc.font = SRH_FONT if own=="SRH" else RCB_FONT   # re-apply after fill
    for c in range(1,11): ws.cell(r,c).border = bdr
    ws.row_dimensions[r].height = 35

ws.freeze_panes = 'A6'
ws.auto_filter.ref = f"A5:J{len(T)+5}"

# ============================================================
# SHEET 2 — TEAM OWNERSHIP
# ============================================================
ws_t = wb.create_sheet("Team Ownership")
ws_t.merge_cells('A1:E1')
ws_t['A1'] = "TEAM OWNERSHIP — SRH (UK, Operations) vs RCB (Bangalore, Technical)"
ws_t['A1'].font = Font(bold=True, size=14, color="1B2A4A")
ws_t['A1'].alignment = Alignment(horizontal='center')

srh_n = sum(1 for t in T if t[5]=="SRH")
rcb_n = sum(1 for t in T if t[5]=="RCB")
tot = len(T)

for col,h in enumerate(["Team","Location","Tasks","% Share","Scope"],1):
    c = ws_t.cell(3,col,h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.border=bdr
ws_t.column_dimensions['A'].width=15; ws_t.column_dimensions['B'].width=18
ws_t.column_dimensions['C'].width=10; ws_t.column_dimensions['D'].width=10; ws_t.column_dimensions['E'].width=70

rows = [
    ("🔵 SRH","United Kingdom 🇬🇧",srh_n,f"{srh_n/tot*100:.0f}%","Legal, Compliance, Insurance, Finance, Recruitment, Marketing, Content, Operations, Support, Physio Onboarding"),
    ("🟣 RCB","Bangalore, India 🇮🇳",rcb_n,f"{rcb_n/tot*100:.0f}%","Infrastructure, Backend, Frontend, Payments, Testing/QA, Design, DevOps, Security, Monitoring, Mobile App"),
    ("TOTAL","",tot,"100%",""),
]
for i,(team,loc,cnt,pct,scope) in enumerate(rows):
    r = i+4
    ws_t.cell(r,1,team).font=Font(bold=True); ws_t.cell(r,2,loc); ws_t.cell(r,3,cnt).alignment=Alignment(horizontal='center')
    ws_t.cell(r,4,pct).alignment=Alignment(horizontal='center'); ws_t.cell(r,5,scope).alignment=Alignment(wrap_text=True)
    for c in range(1,6): ws_t.cell(r,c).border=bdr
    ws_t.row_dimensions[r].height=30

# ── Payment section ──
ws_t.cell(8,1,"💰 PAYMENT SCHEDULE: SRH → RCB").font=Font(bold=True,size=13,color="D32F2F")
ws_t.merge_cells('A8:E8')

for col,h in enumerate(["Date","Instalment","Amount (₹)","Cumulative (₹)","Status"],1):
    c=ws_t.cell(9,col,h); c.font=HEADER_FONT
    c.fill=PatternFill(start_color="D32F2F",end_color="D32F2F",fill_type="solid"); c.border=bdr

pays = [
    ("4 March 2026 (Wed)","1st Instalment","₹50,000","₹50,000","Pending"),
    ("5 March 2026 (Thu)","2nd Instalment","₹50,000","₹1,00,000","Pending"),
    ("6 March 2026 (Fri)","3rd & FINAL Instalment","₹50,000","₹1,50,000 ✅","Pending"),
    ("","TOTAL","₹1,50,000","",""),
]
for i,(d,inst,amt,cum,st) in enumerate(pays):
    r=i+10
    ws_t.cell(r,1,d).border=bdr; ws_t.cell(r,2,inst).border=bdr
    ws_t.cell(r,3,amt).font=Font(bold=True,size=11); ws_t.cell(r,3).border=bdr
    ws_t.cell(r,4,cum).border=bdr; ws_t.cell(r,5,st).border=bdr
    if i==3:
        for c in range(1,6):
            ws_t.cell(r,c).font=Font(bold=True,size=12,color="D32F2F")
            ws_t.cell(r,c).fill=PatternFill(start_color="FFCDD2",end_color="FFCDD2",fill_type="solid")

ws_t.cell(15,1,"Notes:").font=Font(bold=True)
notes = [
    "• SRH pays RCB ₹1,50,000 total within first 3 days (4-6 March). 3 × ₹50,000 instalments.",
    "• Payment via bank transfer / UPI. Collect receipt for each instalment.",
    "• Covers RCB's full platform build through launch (30 March).",
    "• Post-launch technical work (Month 2+) to be agreed as separate engagement.",
    "• RCB technical work begins Day 1 but first payment is on Day 3 (good faith start).",
]
for i,n in enumerate(notes): ws_t.cell(16+i,1,n)

# ── Category breakdown per team ──
r_start = 23
ws_t.cell(r_start,1,"📋 SRH — Tasks by Category").font=Font(bold=True,size=12,color="1565C0")
ws_t.merge_cells(f'A{r_start}:C{r_start}')
r_start+=1
ws_t.cell(r_start,1,"Category").font=Font(bold=True)
ws_t.cell(r_start,2,"Tasks").font=Font(bold=True)
for c in (1,2):
    ws_t.cell(r_start,c).fill=PatternFill(start_color="E3F2FD",end_color="E3F2FD",fill_type="solid")
    ws_t.cell(r_start,c).border=bdr

srh_cats={}; rcb_cats={}
for t in T:
    d = srh_cats if t[5]=="SRH" else rcb_cats
    d[t[2]] = d.get(t[2],0)+1
for cat,cnt in sorted(srh_cats.items(),key=lambda x:-x[1]):
    r_start+=1
    ws_t.cell(r_start,1,cat).border=bdr
    ws_t.cell(r_start,2,cnt).border=bdr; ws_t.cell(r_start,2).alignment=Alignment(horizontal='center')

r_start+=2
ws_t.cell(r_start,1,"📋 RCB — Tasks by Category").font=Font(bold=True,size=12,color="7B1FA2")
ws_t.merge_cells(f'A{r_start}:C{r_start}')
r_start+=1
ws_t.cell(r_start,1,"Category").font=Font(bold=True)
ws_t.cell(r_start,2,"Tasks").font=Font(bold=True)
for c in (1,2):
    ws_t.cell(r_start,c).fill=PatternFill(start_color="F3E5F5",end_color="F3E5F5",fill_type="solid")
    ws_t.cell(r_start,c).border=bdr
for cat,cnt in sorted(rcb_cats.items(),key=lambda x:-x[1]):
    r_start+=1
    ws_t.cell(r_start,1,cat).border=bdr
    ws_t.cell(r_start,2,cnt).border=bdr; ws_t.cell(r_start,2).alignment=Alignment(horizontal='center')


# ============================================================
# SHEET 3 — GANTT OVERVIEW (Launch window only: 2-30 Mar)
# ============================================================
ws2 = wb.create_sheet("Gantt Overview")
ws2.merge_cells('A1:AE1')
ws2['A1'] = "GANTT OVERVIEW — 2 March → 30 March 2026 (28 days to launch)"
ws2['A1'].font = Font(bold=True, size=14, color="1B2A4A")
ws2['A1'].alignment = Alignment(horizontal='center')

ws2.cell(3,1,"Category").font=HEADER_FONT; ws2.cell(3,1).fill=HEADER_FILL
ws2.cell(3,2,"Owner").font=HEADER_FONT; ws2.cell(3,2).fill=HEADER_FILL
ws2.column_dimensions['A'].width=28; ws2.column_dimensions['B'].width=8

sd = date(2026,3,2)
for i in range(29):
    d=sd+timedelta(days=i); col=i+3
    c=ws2.cell(3,col,d.strftime("%d\n%a"))
    c.font=Font(bold=True,color="FFFFFF",size=7); c.fill=HEADER_FILL
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bdr
    ws2.column_dimensions[get_column_letter(col)].width=5.5
ws2.row_dimensions[3].height=32

GANTT_SRH = PatternFill(start_color="42A5F5",end_color="42A5F5",fill_type="solid")
GANTT_RCB = PatternFill(start_color="AB47BC",end_color="AB47BC",fill_type="solid")
LAUNCH_F  = PatternFill(start_color="F44336",end_color="F44336",fill_type="solid")
PAY_F     = PatternFill(start_color="FF5722",end_color="FF5722",fill_type="solid")

# only pre-launch categories
gantt_cats = [
    "Legal & Company Formation","Regulatory & Compliance","Finance & Banking",
    "💰 Payment: SRH → RCB","Insurance","Design & Branding",
    "Technical - Infrastructure","Technical - Backend","Technical - Frontend",
    "Technical - Payments","Technical - Testing & QA","Content & Copywriting",
    "Physio Recruitment","Marketing & Pre-Launch","Operations & Support","Launch Day",
]

for ri,cat in enumerate(gantt_cats,4):
    # figure out primary owner for this category
    cat_owners = [t[5] for t in T if t[2]==cat]
    primary = max(set(cat_owners),key=cat_owners.count) if cat_owners else "SRH"
    ws2.cell(ri,1,cat)
    if cat in CAT_CLR: ws2.cell(ri,1).fill=CAT_CLR[cat]
    ws2.cell(ri,1).border=bdr
    ow = ws2.cell(ri,2,primary); ow.alignment=Alignment(horizontal='center')
    ow.font = SRH_FONT if primary=="SRH" else RCB_FONT; ow.border=bdr
    ws2.row_dimensions[ri].height=20

    for t in T:
        if t[2]==cat:
            off=(t[0]-sd).days
            if 0<=off<29:
                cl=off+3; cell=ws2.cell(ri,cl)
                if "Launch" in cat: cell.fill=LAUNCH_F
                elif "Payment" in cat: cell.fill=PAY_F
                elif t[5]=="SRH": cell.fill=GANTT_SRH
                else: cell.fill=GANTT_RCB
                cell.border=bdr

ws2.freeze_panes='C4'


# ============================================================
# SHEET 4 — BUDGET
# ============================================================
ws3 = wb.create_sheet("Budget Estimate")
ws3.merge_cells('A1:F1')
ws3['A1'] = "BUDGET ESTIMATE"
ws3['A1'].font = Font(bold=True, size=14, color="1B2A4A")
ws3['A1'].alignment = Alignment(horizontal='center')

for col,h in enumerate(["Category","Item","Cost (£)","Frequency","Paid By","Notes"],1):
    c=ws3.cell(3,col,h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.border=bdr
ws3.column_dimensions['A'].width=22; ws3.column_dimensions['B'].width=42
ws3.column_dimensions['C'].width=11; ws3.column_dimensions['D'].width=14
ws3.column_dimensions['E'].width=10; ws3.column_dimensions['F'].width=45

budget = [
    ("💰 RCB Payment","Technical dev fee (₹1,50,000)",1450,"One-off","SRH","₹1.5L ≈ ~£1,450. 3 instalments 4-6 Mar"),
    ("Company","Companies House registration",12,"One-off","SRH",""),
    ("Company","Registered office address",50,"Annual","SRH","Virtual office if needed"),
    ("Legal","Solicitor review (T&Cs, Privacy)",500,"One-off","SRH","Template + review £300-700"),
    ("Legal","ICO registration",40,"Annual","SRH","Required by law"),
    ("Domains","Domain (.co.uk + .com)",20,"Annual","SRH",""),
    ("Hosting","Vercel Pro",20,"Monthly","SRH","Free tier may suffice initially"),
    ("Database","Supabase Pro",25,"Monthly","SRH","Free tier to start"),
    ("Payments","Stripe fees",0,"Per txn","—","1.4% + 20p per UK card"),
    ("Email","SendGrid / Resend",0,"Monthly","SRH","Free tier: 100/day"),
    ("Email Mktg","Mailchimp / Brevo",0,"Monthly","SRH","Free ≤500 contacts"),
    ("Insurance","Public Liability",300,"Annual","SRH","£200-500 range"),
    ("Insurance","Professional Indemnity",400,"Annual","SRH","£300-600 range"),
    ("Design","Logo (Fiverr/99designs)",100,"One-off","RCB","Expense or included in fee"),
    ("Marketing","Google Ads (month 1)",500,"Monthly","SRH","£15-20/day"),
    ("Marketing","Social media ads",300,"Monthly","SRH","Facebook / Instagram"),
    ("Monitoring","Sentry",0,"Monthly","RCB","Free tier"),
    ("Monitoring","UptimeRobot",0,"Monthly","RCB","Free tier"),
    ("Support","Tawk.to live chat",0,"Monthly","SRH","Free"),
    ("SMS","Twilio (optional)",50,"Monthly","SRH","~3p per SMS"),
    ("Misc","Contingency buffer",200,"One-off","SRH",""),
]

total=0
for i,(cat,item,cost,freq,paid,note) in enumerate(budget):
    r=i+4
    ws3.cell(r,1,cat).border=bdr; ws3.cell(r,2,item).border=bdr
    cc=ws3.cell(r,3,cost); cc.number_format='£#,##0'; cc.border=bdr
    ws3.cell(r,4,freq).border=bdr
    ws3.cell(r,5,paid).border=bdr; ws3.cell(r,5).alignment=Alignment(horizontal='center')
    ws3.cell(r,6,note).border=bdr
    total+=cost
    if "RCB Payment" in cat:
        for c in range(1,7):
            ws3.cell(r,c).fill=PatternFill(start_color="FFCDD2",end_color="FFCDD2",fill_type="solid")
            ws3.cell(r,c).font=Font(bold=True)

tr=len(budget)+4
ws3.cell(tr,2,"ESTIMATED TOTAL").font=Font(bold=True,size=12)
ws3.cell(tr,2).border=bdr
ws3.cell(tr,3,total).font=Font(bold=True,size=12)
ws3.cell(tr,3).number_format='£#,##0'; ws3.cell(tr,3).border=bdr


# ============================================================
# SHEET 5 — KEY RESOURCES
# ============================================================
ws4 = wb.create_sheet("Key Resources")
ws4.merge_cells('A1:D1')
ws4['A1'] = "KEY RESOURCES & LINKS"
ws4['A1'].font = Font(bold=True, size=14, color="1B2A4A")
ws4['A1'].alignment = Alignment(horizontal='center')

for col,h in enumerate(["Category","Resource","URL","Notes"],1):
    c=ws4.cell(3,col,h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.border=bdr
ws4.column_dimensions['A'].width=16; ws4.column_dimensions['B'].width=30
ws4.column_dimensions['C'].width=40; ws4.column_dimensions['D'].width=35

res = [
    ("Legal","Companies House","companieshouse.gov.uk","Registration"),
    ("Legal","ICO","ico.org.uk","Data protection"),
    ("Legal","HMRC","gov.uk/register-for-self-assessment","Tax"),
    ("Regulatory","HCPC Register","hcpc-uk.org/check-the-register/","Verify physios"),
    ("Regulatory","CSP","csp.org.uk","Chartered Society"),
    ("Banking","Tide","tide.co","Fast business banking"),
    ("Insurance","Hiscox","hiscox.co.uk","Business insurance"),
    ("Insurance","Simply Business","simplybusiness.co.uk","Compare"),
    ("Technical","Vercel","vercel.com","Hosting"),
    ("Technical","Supabase","supabase.com","DB + Auth"),
    ("Technical","Stripe Connect","stripe.com/connect","Payments"),
    ("Technical","SendGrid","sendgrid.com","Email"),
    ("Technical","Sentry","sentry.io","Error monitoring"),
    ("Technical","postcodes.io","postcodes.io","UK postcode API"),
    ("Marketing","Google Ads","ads.google.com","PPC"),
    ("Marketing","Meta Business","business.facebook.com","Social ads"),
    ("Marketing","Mailchimp","mailchimp.com","Email marketing"),
    ("Recruitment","CSP Jobs","csp.org.uk/jobs","Physio jobs"),
    ("Recruitment","Indeed UK","indeed.co.uk","Jobs"),
    ("Recruitment","LinkedIn","linkedin.com","Professional"),
    ("Funding","Innovate UK","innovateuk.ukri.org","Grants"),
    ("Compliance","CQC","cqc.org.uk","Care Quality Commission"),
]
for i,r in enumerate(res):
    for col,v in enumerate(r,1): ws4.cell(i+4,col,v).border=bdr


# ============================================================
# SHEET 6 — RISK REGISTER
# ============================================================
ws5 = wb.create_sheet("Risk Register")
ws5.merge_cells('A1:F1')
ws5['A1'] = "RISK REGISTER"
ws5['A1'].font = Font(bold=True, size=14, color="1B2A4A"); ws5['A1'].alignment = Alignment(horizontal='center')

for col,h in enumerate(["#","Risk","Impact","Likelihood","Mitigation","Owner"],1):
    c=ws5.cell(3,col,h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.border=bdr
ws5.column_dimensions['A'].width=4; ws5.column_dimensions['B'].width=35
ws5.column_dimensions['C'].width=10; ws5.column_dimensions['D'].width=12
ws5.column_dimensions['E'].width=58; ws5.column_dimensions['F'].width=8

risks = [
    ("Not enough physios at launch","HIGH","MEDIUM","Recruit early. 10-15 min. Focus 1-2 areas. Direct outreach","SRH"),
    ("Stripe Connect approval delayed","HIGH","LOW","Apply Day 5. Manual invoicing backup. Chase Stripe support","SRH"),
    ("HCPC verification issues","HIGH","LOW","Manual backup. Cross-ref HCPC online register directly","RCB"),
    ("Low patient demand at launch","MEDIUM","MEDIUM","Waitlist, Google Ads, social media, local press, GP leaflets","SRH"),
    ("Technical bugs on launch day","HIGH","MEDIUM","Extensive QA, staging env, Sentry. RCB on-call all day (IST evening)","RCB"),
    ("Legal / compliance gap","HIGH","LOW","Early solicitor review. HCPC, ICO, GDPR. Safeguarding policy","SRH"),
    ("Physio no-shows / poor quality","HIGH","LOW","Vetting, reviews, agreement standards. Suspension policy","SRH"),
    ("Payment disputes / chargebacks","MEDIUM","MEDIUM","Clear T&Cs, cancellation policy, Stripe Radar, dispute process","RCB"),
    ("Data breach / security incident","HIGH","LOW","HTTPS, encryption, RLS, backups, GDPR compliance, ICO plan","RCB"),
    ("RCB payment delayed by SRH","HIGH","LOW","Payment split 3 × ₹50k over days 3-5. Blocks dev if delayed","SRH"),
    ("Timezone friction (UK ↔ Bangalore)","MEDIUM","MEDIUM","Daily standup overlap (1-3pm UK / 6:30-8:30pm IST). Async Slack","SRH"),
    ("IR35 / contractor status risk","HIGH","LOW","Legal advice on IR35. HMRC CEST tool. Compliant contracts","SRH"),
]
for i,(risk,imp,lik,mit,own) in enumerate(risks,1):
    r=i+3
    ws5.cell(r,1,i).border=bdr
    ws5.cell(r,2,risk).border=bdr
    c_imp=ws5.cell(r,3,imp); c_imp.border=bdr
    if imp=="HIGH": c_imp.font=Font(bold=True,color="D32F2F")
    elif imp=="MEDIUM": c_imp.font=Font(bold=True,color="E65100")
    c_lik=ws5.cell(r,4,lik); c_lik.border=bdr
    if lik=="MEDIUM": c_lik.font=Font(bold=True,color="E65100")
    ws5.cell(r,5,mit).border=bdr; ws5.cell(r,5).alignment=Alignment(wrap_text=True)
    ow=ws5.cell(r,6,own); ow.border=bdr; ow.alignment=Alignment(horizontal='center')
    ow.font=SRH_FONT if own=="SRH" else RCB_FONT
    ws5.row_dimensions[r].height=35


# ============================================================
# SHEET 7 — LAUNCH DAY CHECKLIST (30 March)
# ============================================================
ws6 = wb.create_sheet("Launch Day Checklist")
ws6.merge_cells('A1:D1')
ws6['A1'] = "🚀 LAUNCH DAY — 30 MARCH 2026 (Monday)"
ws6['A1'].font = Font(bold=True, size=14, color="D32F2F"); ws6['A1'].alignment = Alignment(horizontal='center')

ws6.merge_cells('A2:D2')
ws6['A2'] = "RCB works IST (overlap with UK: 1:30pm-8:30pm IST = 8am-3pm UK)  |  SRH works UK time"
ws6['A2'].font = Font(size=10, italic=True, color="666666"); ws6['A2'].alignment = Alignment(horizontal='center')

for col,h in enumerate(["Time (UK)","Task","Owner","Done?"],1):
    c=ws6.cell(4,col,h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.border=bdr
ws6.column_dimensions['A'].width=12; ws6.column_dimensions['B'].width=58
ws6.column_dimensions['C'].width=8; ws6.column_dimensions['D'].width=8

checks = [
    ("06:30","Final system health check — all services green","RCB",""),
    ("07:00","Verify Stripe live mode active","RCB",""),
    ("07:00","Verify email delivery (send test)","RCB",""),
    ("07:30","Remove invite-only / beta flags","RCB",""),
    ("07:30","Verify site accessible at live domain","RCB",""),
    ("08:00","Post launch announcement — LinkedIn","SRH",""),
    ("08:00","Post launch announcement — Instagram","SRH",""),
    ("08:00","Post launch announcement — Facebook","SRH",""),
    ("08:00","Post launch announcement — Twitter/X","SRH",""),
    ("08:30","Send launch email to mailing list","SRH",""),
    ("08:30","Activate Google Ads campaigns","SRH",""),
    ("09:00","Monitor first sign-ups and bookings","SRH",""),
    ("09:00","Check Sentry error dashboard","RCB",""),
    ("10:00","Verify all physios showing in search","SRH",""),
    ("10:00","Test real booking flow on production","RCB",""),
    ("12:00","Midday metrics: sign-ups, bookings, errors","SRH",""),
    ("12:00","Respond to all support enquiries","SRH",""),
    ("14:00","Post afternoon social media update","SRH",""),
    ("14:30","Check payment processing (RCB overlap ends ~3pm UK)","RCB",""),
    ("17:00","End of day metrics review","SRH",""),
    ("17:00","Send thank-you to all physios","SRH",""),
    ("18:00","Evening social media engagement","SRH",""),
    ("21:00","Final system check (RCB morning IST)","RCB",""),
]
for i,(time,task,own,done) in enumerate(checks):
    r=i+5
    ws6.cell(r,1,time).border=bdr; ws6.cell(r,1).font=Font(bold=True); ws6.cell(r,1).alignment=Alignment(horizontal='center')
    ws6.cell(r,2,task).border=bdr
    ow=ws6.cell(r,3,own); ow.border=bdr; ow.alignment=Alignment(horizontal='center')
    ow.font=SRH_FONT if own=="SRH" else RCB_FONT
    ws6.cell(r,4,done).border=bdr
    ws6.row_dimensions[r].height=22


# ============================================================
# SAVE
# ============================================================
out = "/Users/safestorage/Desktop/APPLICATIONS/physio-connect/PhysioConnect_Launch_Plan.xlsx"
wb.save(out)
print(f"✅ Saved: {out}")
print(f"📊 Total tasks: {len(T)}")
print(f"🔵 SRH: {srh_n}  |  🟣 RCB: {rcb_n}")
print(f"📑 Sheets: {wb.sheetnames}")

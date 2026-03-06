import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Bug Reports"

# Define colors
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define headers
headers = [
    "S.No",
    "Date",
    "Reported By",
    "Module/Group",
    "Tab/Section",
    "Field Name",
    "Issue/Suggestion",
    "Notes/Comments",
    "Priority",
    "Severity",
    "Status",
    "Assigned To",
    "Resolution Date",
    "Resolution Notes"
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Set column widths
column_widths = [8, 12, 15, 20, 25, 20, 40, 35, 12, 12, 12, 15, 15, 30]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Define module data structure
modules = {
    "Dashboard": ["Main Dashboard", "Statistics Cards", "Revenue Charts", "Recent Activities", "Quick Actions"],
    "Activities": ["All Activities", "Follow-ups", "Tasks", "Create Activity", "Mark Complete", "Calendar View"],
    "Sales": ["Leads", "Contacts", "Accounts", "Deals"],
    "Inventory": ["Products", "Stock Entries", "Quotations", "Sales Orders", "Invoices"],
    "Services": ["Installations", "AMC", "Complaints"],
    "Support": ["Support Center", "Support Tickets"],
    "Analytics": ["MIS Reports", "Sales Reports", "Service Reports", "Inventory Reports"],
    "Resources": ["Doc Library", "Documents"],
    "Admin": ["Company Settings", "User Management", "Role Management", "Profile Settings", "Company Logo"]
}

# Create a reference sheet for dropdowns
ref_sheet = wb.create_sheet("Reference")
ref_sheet.sheet_state = 'hidden'

# Write modules to reference sheet
ref_sheet['A1'] = "Modules"
ref_sheet['A1'].font = Font(bold=True)
module_list = list(modules.keys())
for i, module in enumerate(module_list, 2):
    ref_sheet[f'A{i}'] = module

# Write tabs for each module and create named ranges
col = 2
col_mapping = {}
for module, tabs in modules.items():
    # Store column letter for this module
    col_letter = get_column_letter(col)
    col_mapping[module] = col_letter

    # Write header
    header_name = f"{module}_Tabs"
    ref_sheet.cell(row=1, column=col).value = header_name
    ref_sheet.cell(row=1, column=col).font = Font(bold=True)

    # Write tabs
    max_row = 1
    for i, tab in enumerate(tabs, 2):
        ref_sheet.cell(row=i, column=col).value = tab
        max_row = i

    # Create a named range for this module's tabs
    from openpyxl.workbook.defined_name import DefinedName
    # Named range should reference the cells containing the tabs
    range_ref = f"Reference!${col_letter}$2:${col_letter}${max_row}"
    wb.defined_names[header_name] = DefinedName(header_name, attr_text=range_ref)

    col += 1

# Priority options
ref_sheet['M1'] = "Priority"
ref_sheet['M1'].font = Font(bold=True)
priorities = ["Critical", "High", "Medium", "Low"]
for i, priority in enumerate(priorities, 2):
    ref_sheet[f'M{i}'] = priority

# Severity options
ref_sheet['N1'] = "Severity"
ref_sheet['N1'].font = Font(bold=True)
severities = ["Blocker", "Critical", "Major", "Minor", "Trivial"]
for i, severity in enumerate(severities, 2):
    ref_sheet[f'N{i}'] = severity

# Status options
ref_sheet['O1'] = "Status"
ref_sheet['O1'].font = Font(bold=True)
statuses = ["Open", "In Progress", "Fixed", "Closed", "Rejected", "Duplicate", "Need More Info"]
for i, status in enumerate(statuses, 2):
    ref_sheet[f'O{i}'] = status

# Add data validations to main sheet
# Module dropdown (Column D)
module_dv = DataValidation(type="list", formula1=f"=Reference!$A$2:$A${len(module_list)+1}", allow_blank=False)
module_dv.error = 'Please select a module from the list'
module_dv.errorTitle = 'Invalid Module'
ws.add_data_validation(module_dv)
module_dv.add(f'D2:D1000')

# Tab/Section dropdown (Column E) - Conditional based on Module
# We'll add individual dropdowns for each row that reference the correct column based on module selected
# For now, add a general note that users should type based on module selected
# The conditional dropdown will be added using formulas

# Priority dropdown (Column I)
priority_dv = DataValidation(type="list", formula1=f"=Reference!$M$2:$M${len(priorities)+1}", allow_blank=False)
priority_dv.error = 'Please select a priority from the list'
priority_dv.errorTitle = 'Invalid Priority'
ws.add_data_validation(priority_dv)
priority_dv.add(f'I2:I1000')

# Severity dropdown (Column J)
severity_dv = DataValidation(type="list", formula1=f"=Reference!$N$2:$N${len(severities)+1}", allow_blank=False)
severity_dv.error = 'Please select a severity from the list'
severity_dv.errorTitle = 'Invalid Severity'
ws.add_data_validation(severity_dv)
severity_dv.add(f'J2:J1000')

# Status dropdown (Column K)
status_dv = DataValidation(type="list", formula1=f"=Reference!$O$2:$O${len(statuses)+1}", allow_blank=False)
status_dv.error = 'Please select a status from the list'
status_dv.errorTitle = 'Invalid Status'
ws.add_data_validation(status_dv)
status_dv.add(f'K2:K1000')

# Add conditional dropdowns for Tab/Section based on Module selection
# This uses INDIRECT formula to create dynamic dependent dropdowns
for row in range(2, 1001):
    cell = ws.cell(row=row, column=5)  # Column E (Tab/Section)
    # Create a conditional dropdown using INDIRECT
    tab_dv = DataValidation(type="list",
                           formula1=f'=INDIRECT(SUBSTITUTE(D{row}," ","_")&"_Tabs")',
                           allow_blank=True)
    tab_dv.error = 'Please select a valid tab for the chosen module'
    tab_dv.errorTitle = 'Invalid Tab'
    ws.add_data_validation(tab_dv)
    tab_dv.add(f'E{row}')

# Add sample data row with formulas
ws['A2'] = 1
ws['B2'] = '=TODAY()'
ws['C2'] = 'Tester Name'
ws['D2'] = 'Dashboard'
ws['E2'] = 'Main Dashboard'
ws['F2'] = 'Revenue Card'
ws['G2'] = 'Revenue not displaying correctly'
ws['H2'] = 'Format issue with currency display'
ws['I2'] = 'Medium'
ws['J2'] = 'Minor'
ws['K2'] = 'Open'

# Create Instructions sheet
inst_sheet = wb.create_sheet("Instructions")
inst_sheet['A1'] = "CRM BUG TRACKING TEMPLATE - INSTRUCTIONS"
inst_sheet['A1'].font = Font(bold=True, size=14, color="1F4E78")
inst_sheet.merge_cells('A1:D1')

instructions = [
    "",
    "HOW TO USE THIS TEMPLATE:",
    "",
    "1. S.No: Sequential number for each bug (1, 2, 3...)",
    "",
    "2. Date: Auto-filled with today's date or enter manually",
    "",
    "3. Reported By: Enter your name",
    "",
    "4. Module/Group: Select from dropdown",
    "   • Dashboard - Main dashboard and statistics",
    "   • Activities - Follow-ups, tasks, activities",
    "   • Sales - Leads, Contacts, Accounts, Deals",
    "   • Inventory - Products, Stock Entries, Quotations, Sales Orders, Invoices",
    "   • Services - Installations, AMC, Complaints",
    "   • Support - Support Center and Tickets",
    "   • Analytics - MIS Reports",
    "   • Resources - Document Library",
    "   • Admin - Settings, User Management",
    "",
    "5. Tab/Section: Enter the specific tab or section",
    "   Based on module selected:",
    "   • Dashboard: Main Dashboard, Statistics Cards, Revenue Charts, etc.",
    "   • Activities: All Activities, Follow-ups, Tasks, etc.",
    "   • Sales: Leads, Contacts, Accounts, Deals",
    "   • Inventory: Products, Stock Entries, Quotations, Sales Orders, Invoices",
    "   • Services: Installations, AMC, Complaints",
    "   • Support: Support Center, Support Tickets",
    "   • Analytics: MIS Reports, Sales Reports, Service Reports, Inventory Reports",
    "   • Resources: Doc Library, Documents",
    "   • Admin: Company Settings, User Management, Role Management, Profile Settings",
    "",
    "6. Field Name: Specific field where issue occurs (if applicable)",
    "   Examples: 'Product Name', 'Email Field', 'Submit Button', 'Date Picker'",
    "",
    "7. Issue/Suggestion: Clear description of the issue or enhancement request",
    "   • Be specific and detailed",
    "   • Include steps to reproduce if it's a bug",
    "   • Include expected vs actual behavior",
    "",
    "8. Notes/Comments: Additional information, screenshots path, or context",
    "",
    "9. Priority: Select from dropdown",
    "   • Critical - System crash, data loss, security issue",
    "   • High - Major feature broken, blocking work",
    "   • Medium - Feature works but has issues",
    "   • Low - Minor cosmetic issue, nice-to-have",
    "",
    "10. Severity: Select from dropdown",
    "   • Blocker - Prevents testing/usage completely",
    "   • Critical - Major functionality broken",
    "   • Major - Important feature affected",
    "   • Minor - Small issue, workaround available",
    "   • Trivial - Cosmetic issue only",
    "",
    "11. Status: Select from dropdown",
    "   • Open - Newly reported",
    "   • In Progress - Being worked on",
    "   • Fixed - Developer has fixed",
    "   • Closed - Verified and closed",
    "   • Rejected - Not a bug / Won't fix",
    "   • Duplicate - Already reported",
    "   • Need More Info - Requires additional details",
    "",
    "12. Assigned To: Developer/Team member name (filled by admin)",
    "",
    "13. Resolution Date: Date when issue was resolved",
    "",
    "14. Resolution Notes: How the issue was fixed",
    "",
    "",
    "TESTING CHECKLIST:",
    "☐ Test all CRUD operations (Create, Read, Update, Delete)",
    "☐ Test form validations (required fields, email format, etc.)",
    "☐ Test search and filter functionality",
    "☐ Test on different browsers (Chrome, Safari, Firefox)",
    "☐ Test on different devices (Desktop, Tablet, Mobile)",
    "☐ Test with different user roles (Admin, Manager, Sales, etc.)",
    "☐ Test data consistency across modules",
    "☐ Test error handling and error messages",
    "☐ Test loading states and performance",
    "☐ Test navigation and breadcrumbs",
    "",
    "",
    "MODULE-WISE TESTING AREAS:",
    "",
    "DASHBOARD:",
    "• Statistics cards displaying correct data",
    "• Charts loading and interactive",
    "• Recent activities showing correctly",
    "• Quick actions working",
    "",
    "ACTIVITIES:",
    "• Create new activity/follow-up/task",
    "• Edit existing activities",
    "• Mark as complete",
    "• Filter by status, type, date",
    "• Calendar view working",
    "",
    "SALES - LEADS:",
    "• Create/Edit/Delete lead",
    "• Convert lead to deal",
    "• Lead status updates",
    "• Search and filter",
    "",
    "SALES - CONTACTS:",
    "• Create/Edit/Delete contact",
    "• Link to account",
    "• Contact details display",
    "",
    "SALES - ACCOUNTS:",
    "• Create/Edit/Delete account",
    "• Related contacts display",
    "• Account details",
    "",
    "SALES - DEALS:",
    "• Create/Edit/Delete deal",
    "• Multiple products selection",
    "• Deal amount calculation",
    "• Deal stage management",
    "",
    "INVENTORY - PRODUCTS:",
    "• Create/Edit/Delete product",
    "• Stock quantity display",
    "• Category filtering",
    "• Product search",
    "",
    "INVENTORY - STOCK ENTRIES:",
    "• Create stock inward entry",
    "• Create stock outward entry",
    "• Entry approval workflow",
    "• Stock validation (prevent negative stock)",
    "• Bin location update",
    "• Stock summary accuracy",
    "• Product details tooltip",
    "",
    "INVENTORY - QUOTATIONS:",
    "• Create/Edit/Delete quotation",
    "• Multiple products",
    "• Total calculation",
    "• Status management",
    "",
    "SERVICES - INSTALLATIONS:",
    "• Create/Edit/Delete installation",
    "• Assign technician",
    "• Schedule date",
    "• Status tracking",
    "",
    "SERVICES - AMC:",
    "• Create/Edit/Delete AMC",
    "• Renewal management",
    "• Service schedule",
    "• AMC status",
    "",
    "SERVICES - COMPLAINTS:",
    "• Create/Edit/Delete complaint",
    "• Priority management",
    "• Status tracking",
    "• Solution recording",
    "• Assign technician",
    "",
    "SUPPORT:",
    "• Create support ticket",
    "• Ticket status update",
    "• Assign support agent",
    "• Priority management",
    "",
    "ANALYTICS:",
    "• MIS reports display",
    "• Filter by date range",
    "• Export reports",
    "• Data accuracy",
    "",
    "ADMIN:",
    "• Company settings update",
    "• User management (Add/Edit/Delete users)",
    "• Role assignment",
    "• Company logo upload and display",
    "• Profile settings",
]

for i, instruction in enumerate(instructions, 2):
    inst_sheet[f'A{i}'] = instruction
    if instruction.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '10.', '11.', '12.', '13.', '14.')):
        inst_sheet[f'A{i}'].font = Font(bold=True, color="1F4E78")
    elif instruction in ["HOW TO USE THIS TEMPLATE:", "TESTING CHECKLIST:", "MODULE-WISE TESTING AREAS:"]:
        inst_sheet[f'A{i}'].font = Font(bold=True, size=12, color="1F4E78")
    elif instruction.endswith(':') and not instruction.startswith('☐'):
        inst_sheet[f'A{i}'].font = Font(bold=True, color="C00000")

inst_sheet.column_dimensions['A'].width = 100

# Freeze panes on main sheet
ws.freeze_panes = 'A2'

# Auto-filter on headers
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

# Save the workbook
wb.save('/Users/safestorage/Desktop/crm1 (1)/CRM_Bug_Tracking_Template.xlsx')
print("Excel template created successfully!")
